import openpyxl

input_file = 'CERNOX Calibration Summary_V2.xlsx'
output_file = 'CernoxCalibration_Consolidated.xlsx'

wb = openpyxl.load_workbook(input_file, use_iterators=True)
out_file = openpyxl.Workbook()
sheet = out_file.active
sheet.title = 'data'

for sheet_num, sh in enumerate(wb.worksheets):
    #print 'worksheet title is %s' % sh.title
    ser_num = sh.title

    temperatures = []
    resistances = []

    #cycle through the rows in the sheet
    #cycle through the cells in each row
    #fill the temperature and resistance arrays with the values
    for row in sh.iter_rows():
        for cell_index, cell in enumerate(row):

            if cell.data_type == 's' or cell.value == None:
                #skip the non-data values
                continue

            val = float(cell.value)

            if cell_index == 0:
                temperatures.append(val)
            elif cell_index == 1:
                resistances.append(val)
            else:
                continue
# this is ugly, but it writes columns of data
# skipping 4.22 and 273.15 because not every data set had those points
    rownum = 2
    if sheet_num == 0:
        sheet.cell(row=1, column=2).value = ser_num
        for temperature in temperatures:
            if temperature == 4.22 or temperature == 273.15:
                continue
            sheet.cell(row=rownum,column=1).value = temperature
            rownum += 1
        rownum = 2
        for r, resistance in enumerate(resistances):
            if temperatures[r] == 4.22 or temperatures[r] == 273.15:
                continue
            sheet.cell(row=rownum, column=2).value = resistance
            rownum += 1
    else:
        sheet.cell(row=1, column=sheet_num+2).value = ser_num
        for r, resistance in enumerate(resistances):
            if temperatures[r] == 4.22 or temperatures[r] == 273.15:
                continue
            sheet.cell(row=rownum, column=sheet_num+2).value = resistance
            rownum += 1
out_file.save(filename = output_file)
