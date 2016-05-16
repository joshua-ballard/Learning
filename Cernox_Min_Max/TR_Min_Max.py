import openpyxl

input_file = 'CernoxList.xlsx'
output_file = 'CernoxStats.txt'
#wb = openpyxl.load_workbook(input_file, use_iterators=True)

most_max = 0.0
most_min = 100000000.0

wb = openpyxl.load_workbook(input_file, use_iterators=True)
with open(output_file, 'w') as out_file:

    out_file.write('SN\tMax T\tRes@Max\tMin T\tRes@Min\n')

    for sh in wb.worksheets:
        #print 'worksheet title is %s' % sh.title
        ser_num = sh.title

        temperatures = []
        resistances = []

        #cycle through the rows in the sheet
        #cycle through the cells in each row
        #fill the temperature and resistance arrays with the values
        for row in sh.iter_rows():
            for cell_index, cell in enumerate(row):

                if cell.data_type == 's' or cell.value == None:
                    #skip the non-data values
                    continue

                val = round(float(cell.value), 5)

                if cell_index == 0:
                    temperatures.append(val)
                elif cell_index == 1:
                    resistances.append(val)
                else:
                    continue

        min_temp = temperatures[0]
        # print min_temp, type(min_temp)
        max_temp = temperatures[-1]
        # print max_temp, type(max_temp)
        min_resistance = resistances[-1]
        # print min_resistance, type(min_resistance)
        max_resistance = resistances[0]
        # print max_resistance, type(max_resistance)

        line_to_write_list = [ser_num, max_temp, min_resistance, min_temp, max_resistance]

        line_to_write = '\t'.join([str(item) for item in line_to_write_list])

        out_file.write(line_to_write+'\n')

        if min_resistance <= most_min:
            most_min = min_resistance

        if max_resistance >= most_max:
            most_max = max_resistance

    out_file.write("Max Resistance = %f \n" % most_max)
    out_file.write("Min Resistance = %f \n" % most_min)

    print "Max Resistance = %f" % most_max
    print "Min Resistance = %f" % most_min
